"""
process_airlab_energy.py — AirLab CMU 无人机飞行能耗数据清洗脚本

将 data/raw/airlab_energy/data/{N}/processed.csv 清洗并合并到:
  - data/processed/airlab_energy/flights_summary.csv  (每次飞行的汇总统计)
  - data/processed/airlab_energy/flights_detail.csv    (全部时序数据，带 flight_id)

来源: CMU AirLab — DJI Matrice 100, 187次飞行
原始字段: time, airspeed, vertspd, psi, aoa, theta, diffalt, density, payload, power, airspeed_x, airspeed_y
"""

import csv
import os
import logging
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("正在安装 openpyxl...")
    os.system("pip install openpyxl -q")
    import openpyxl

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger("AirLabProcessor")

# 项目根目录
BASE_DIR = Path(__file__).resolve().parent.parent
RAW_DIR = BASE_DIR / "data" / "raw" / "airlab_energy" / "data"
OUTPUT_DIR = BASE_DIR / "data" / "processed" / "airlab_energy"
FLIGHT_SHEET = RAW_DIR / "Flight Sheet.xlsx"

# 时序明细中保留的字段
DETAIL_FIELDS = [
    "flight_id", "time", "airspeed", "vertspd", "diffalt",
    "payload", "power", "density", "airspeed_x", "airspeed_y",
    "psi", "aoa", "theta"
]

# 汇总字段
SUMMARY_FIELDS = [
    "flight_id", "flight_number", "route", "aircraft", "date",
    "payload_kg", "duration_s", "max_altitude_m",
    "avg_airspeed_ms", "max_airspeed_ms",
    "avg_power_w", "max_power_w", "min_power_w",
    "total_energy_wh", "energy_per_second_wh",
    "avg_density", "sample_count", "sample_rate_hz"
]


def load_flight_sheet():
    """从 Flight Sheet.xlsx 读取飞行实验参数"""
    if not FLIGHT_SHEET.exists():
        logger.warning(f"未找到 Flight Sheet: {FLIGHT_SHEET}")
        return {}

    wb = openpyxl.load_workbook(FLIGHT_SHEET, read_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else f"col_{i}" for i, c in enumerate(ws[1])]

    flight_meta = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        fnum = int(row[0])
        meta = {}
        for i, val in enumerate(row):
            if i < len(headers):
                meta[headers[i]] = val
        flight_meta[fnum] = meta

    wb.close()
    logger.info(f"从 Flight Sheet 加载了 {len(flight_meta)} 条飞行元数据")
    return flight_meta


def process_single_flight(flight_dir: Path, flight_number: int, flight_meta: dict):
    """处理单个飞行记录，返回 (summary_row, detail_rows)"""
    csv_file = flight_dir / "processed.csv"
    if not csv_file.exists():
        return None, []

    rows = []
    with open(csv_file, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            try:
                parsed = {
                    "time": float(row["time"]),
                    "airspeed": float(row["airspeed"]),
                    "vertspd": float(row["vertspd"]),
                    "psi": float(row["psi"]),
                    "aoa": float(row["aoa"]),
                    "theta": float(row["theta"]),
                    "diffalt": float(row["diffalt"]),
                    "density": float(row["density"]),
                    "payload": float(row["payload"]),
                    "power": float(row["power"]),
                    "airspeed_x": float(row["airspeed_x"]),
                    "airspeed_y": float(row["airspeed_y"]),
                }
                rows.append(parsed)
            except (ValueError, KeyError):
                continue

    if len(rows) < 2:
        return None, []

    # 构建 flight_id
    flight_id = f"AIRLAB_{flight_number:04d}"

    # 计算汇总统计
    times = [r["time"] for r in rows]
    powers = [r["power"] for r in rows]
    airspeeds = [r["airspeed"] for r in rows]
    alts = [r["diffalt"] for r in rows]
    densities = [r["density"] for r in rows]

    duration = max(times) - min(times)
    avg_dt = duration / (len(rows) - 1) if len(rows) > 1 else 0.1
    sample_rate = round(1.0 / avg_dt, 1) if avg_dt > 0 else 10.0

    # 能耗计算: 梯形积分 ∫ power dt，转换为 Wh
    total_energy_j = 0.0
    for i in range(1, len(rows)):
        dt = rows[i]["time"] - rows[i - 1]["time"]
        avg_p = (rows[i]["power"] + rows[i - 1]["power"]) / 2.0
        total_energy_j += avg_p * dt
    total_energy_wh = total_energy_j / 3600.0

    # 获取元数据
    meta = flight_meta.get(flight_number, {})
    route = meta.get("Route #", "")
    aircraft = meta.get("Aircraft #", "")
    date = meta.get("Date [YYYY-MM-DD]", "")
    if hasattr(date, "strftime"):
        date = date.strftime("%Y-%m-%d")

    summary = {
        "flight_id": flight_id,
        "flight_number": flight_number,
        "route": route if route else "",
        "aircraft": aircraft if aircraft else "",
        "date": str(date) if date else "",
        "payload_kg": rows[0]["payload"],
        "duration_s": round(duration, 2),
        "max_altitude_m": round(max(alts), 2),
        "avg_airspeed_ms": round(sum(airspeeds) / len(airspeeds), 3),
        "max_airspeed_ms": round(max(airspeeds), 3),
        "avg_power_w": round(sum(powers) / len(powers), 2),
        "max_power_w": round(max(powers), 2),
        "min_power_w": round(min(powers), 2),
        "total_energy_wh": round(total_energy_wh, 4),
        "energy_per_second_wh": round(total_energy_wh / duration, 6) if duration > 0 else 0,
        "avg_density": round(sum(densities) / len(densities), 6),
        "sample_count": len(rows),
        "sample_rate_hz": sample_rate,
    }

    # 构建明细行
    detail_rows = []
    for r in rows:
        detail = {"flight_id": flight_id}
        detail.update(r)
        # 精度裁剪
        for k in ["time", "airspeed", "vertspd", "diffalt", "density",
                   "power", "airspeed_x", "airspeed_y", "psi", "aoa", "theta"]:
            if k in detail:
                detail[k] = round(detail[k], 4)
        detail["payload"] = round(detail["payload"], 3)
        detail_rows.append(detail)

    return summary, detail_rows


def main():
    logger.info("=" * 60)
    logger.info("AirLab CMU 飞行能耗数据清洗")
    logger.info("=" * 60)

    # 加载飞行元数据
    flight_meta = load_flight_sheet()

    # 发现所有飞行记录目录
    flight_dirs = []
    for d in sorted(RAW_DIR.iterdir()):
        if d.is_dir() and d.name.isdigit():
            flight_dirs.append((int(d.name), d))

    logger.info(f"发现 {len(flight_dirs)} 个飞行记录目录")

    # 处理所有飞行
    summaries = []
    all_details = []
    skipped = 0

    for flight_num, flight_dir in flight_dirs:
        summary, details = process_single_flight(flight_dir, flight_num, flight_meta)
        if summary is None:
            skipped += 1
            continue
        summaries.append(summary)
        all_details.extend(details)

    logger.info(f"成功处理 {len(summaries)} 次飞行, 跳过 {skipped} 次")
    logger.info(f"明细数据共 {len(all_details)} 行")

    # 输出目录
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # 写入汇总 CSV
    summary_file = OUTPUT_DIR / "flights_summary.csv"
    with open(summary_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=SUMMARY_FIELDS)
        writer.writeheader()
        writer.writerows(summaries)
    logger.info(f"✅ 汇总文件: {summary_file} ({len(summaries)} 行, {summary_file.stat().st_size / 1024:.1f} KB)")

    # 写入明细 CSV
    detail_file = OUTPUT_DIR / "flights_detail.csv"
    with open(detail_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=DETAIL_FIELDS)
        writer.writeheader()
        writer.writerows(all_details)
    size_mb = detail_file.stat().st_size / (1024 * 1024)
    logger.info(f"✅ 明细文件: {detail_file} ({len(all_details)} 行, {size_mb:.2f} MB)")

    # 打印统计摘要
    logger.info("")
    logger.info("=" * 60)
    logger.info("📊 数据统计摘要:")
    logger.info(f"  总飞行次数: {len(summaries)}")
    total_duration = sum(s["duration_s"] for s in summaries)
    logger.info(f"  总飞行时长: {total_duration:.0f} 秒 ({total_duration / 3600:.2f} 小时)")
    total_energy = sum(s["total_energy_wh"] for s in summaries)
    logger.info(f"  总能耗: {total_energy:.2f} Wh")

    payloads = set(s["payload_kg"] for s in summaries)
    logger.info(f"  载荷分布: {sorted(payloads)} kg")

    avg_powers = [s["avg_power_w"] for s in summaries]
    logger.info(f"  平均功率范围: {min(avg_powers):.1f} ~ {max(avg_powers):.1f} W")

    max_alts = [s["max_altitude_m"] for s in summaries]
    logger.info(f"  最大飞行高度范围: {min(max_alts):.1f} ~ {max(max_alts):.1f} m")
    logger.info("=" * 60)


if __name__ == "__main__":
    main()
